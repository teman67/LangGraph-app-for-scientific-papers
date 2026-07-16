"""
Parsers for the four kinds of input this app deals with:
  1. The "input" Excel sheet that lists the fields to extract, together with
     the ontology class each field maps to (columns: # | field name |
     ontology class | ontology definition).
  2. The "output" Excel template whose header row defines the exact column
     names / order that the final table must follow.
  3. Source documents: PDFs or plain text files containing the scientific
     paper(s) to mine for data.
  4. The ontology .ttl file (handled in ontology.py).
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import List

import openpyxl
from pypdf import PdfReader


@dataclass
class FieldSpec:
    index: int
    field_name: str
    ontology_class: str
    ontology_definition: str


@dataclass
class ParsedDocument:
    name: str
    text: str
    truncated_intro: bool = False


def parse_input_schema(file_like) -> List[FieldSpec]:
    """Read the 'LLM input - terms from CTO' style workbook.

    Expected columns (any sheet, first non-empty row is a header):
        # | field_name | related ontology class | related ontology class (definition)
    """
    wb = openpyxl.load_workbook(file_like, data_only=True)
    ws = wb[wb.sheetnames[0]]

    specs: List[FieldSpec] = []
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        cells = list(row) + [None] * (4 - len(row)) if len(row) < 4 else list(row)
        idx, name, ont_class, ont_def = cells[0], cells[1], cells[2], cells[3]

        if idx is None and name is None:
            continue
        if not header_seen:
            # skip the header row itself (idx == '#')
            header_seen = True
            continue
        if idx is None or name is None:
            continue
        try:
            idx_int = int(idx)
        except (ValueError, TypeError):
            continue
        specs.append(
            FieldSpec(
                index=idx_int,
                field_name=str(name).strip(),
                ontology_class=str(ont_class).strip() if ont_class else "",
                ontology_definition=str(ont_def).strip() if ont_def else "",
            )
        )
    specs.sort(key=lambda s: s.index)
    return specs


def parse_output_headers(file_like) -> List[str]:
    """Read the header row of the output template workbook."""
    wb = openpyxl.load_workbook(file_like, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(c).strip() if c is not None else "" for c in row]
    return []


_INTRO_HEADING_RE = re.compile(r"^\s*(\d+\.?\s*)?introduction\s*$", re.IGNORECASE)
# Common section headings that would follow an "Introduction" section in a
# materials-science paper. Used as a stop point when stripping the intro.
_NEXT_HEADING_RE = re.compile(
    r"^\s*(\d+\.?\s*)?"
    r"(materials?(\s+and\s+methods?)?|experimental(\s+(procedure|methods?|details?|setup))?|"
    r"methods?|methodology|results?(\s+and\s+discussion)?|background)\s*$",
    re.IGNORECASE,
)


def strip_introduction(text: str) -> tuple[str, bool]:
    """Best-effort removal of the 'Introduction' section of a paper.

    Looks for a line that is (approximately) just the word 'Introduction'
    and removes everything up to the next recognizable section heading.
    This is a heuristic — it will not be perfect for every paper layout —
    but it keeps obviously irrelevant introductory/background text out of
    the extraction context.
    """
    lines = text.splitlines()
    start_idx = None
    for i, line in enumerate(lines):
        if _INTRO_HEADING_RE.match(line):
            start_idx = i
            break
    if start_idx is None:
        return text, False

    end_idx = len(lines)
    for j in range(start_idx + 1, len(lines)):
        if _NEXT_HEADING_RE.match(lines[j]):
            end_idx = j
            break

    new_lines = lines[:start_idx] + lines[end_idx:]
    return "\n".join(new_lines), True


def read_pdf(file_like) -> str:
    """Extract text page by page, inserting `[[page N]]` markers so the LLM
    can cite an approximate page number for provenance."""
    reader = PdfReader(file_like)
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        pages.append(f"[[page {i}]]\n{page.extract_text() or ''}")
    return "\n".join(pages)


def load_document(name: str, raw_bytes: bytes) -> ParsedDocument:
    lower = name.lower()
    if lower.endswith(".pdf"):
        text = read_pdf(io.BytesIO(raw_bytes))
    else:
        text = raw_bytes.decode("utf-8", errors="ignore")

    cleaned, truncated = strip_introduction(text)
    return ParsedDocument(name=name, text=cleaned, truncated_intro=truncated)


def chunk_text(text: str, max_chars: int = 15000, overlap_chars: int = 800) -> List[str]:
    """Split long text into overlapping chunks on paragraph boundaries where
    possible, so long papers don't get silently truncated before their data
    tables/results sections are reached.
    """
    if len(text) <= max_chars:
        return [text]

    paragraphs = re.split(r"(\n\s*\n)", text)
    units = []
    buf = ""
    for part in paragraphs:
        buf += part
        if part.strip() == "":
            units.append(buf)
            buf = ""
    if buf:
        units.append(buf)
    if not units:
        units = [text]

    chunks: List[str] = []
    current = ""
    for unit in units:
        if len(current) + len(unit) > max_chars and current:
            chunks.append(current)
            current = current[-overlap_chars:] + unit
        else:
            current += unit
    if current:
        chunks.append(current)

    final_chunks: List[str] = []
    for c in chunks:
        if len(c) <= max_chars * 1.5:
            final_chunks.append(c)
        else:
            for i in range(0, len(c), max_chars):
                final_chunks.append(c[i : i + max_chars + overlap_chars])
    return final_chunks
